import os
from datetime import date, datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelXLSXService:
    """Excel export service for generating true .xlsx files"""
    
    def __init__(self, data_dir: str = "../backend/data"):
        self.data_dir = data_dir
        self.export_dir = os.path.join(data_dir, "exports")
        os.makedirs(self.export_dir, exist_ok=True)
        
        # Define styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.data_font = Font(size=11)
        self.data_alignment = Alignment(horizontal="left", vertical="center")
        
        self.currency_alignment = Alignment(horizontal="right", vertical="center")
        
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def export_detailed_sales_report(self, bills: List[Dict], summary_data: Dict) -> str:
        """Create a comprehensive Excel report with detailed bills and summary"""
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            filename = f"detailed_sales_report_{today_str}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sales Report"
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            
            current_row = 1
            
            # === SUMMARY SECTION ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "DAILY SALES SUMMARY"
            ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            # Summary data
            summary_items = [
                ("Date", summary_data.get('date', today_str)),
                ("Total Bills", summary_data.get('total_bills', 0)),
                ("Total Sales", f"{summary_data.get('total_sales', 0):.2f}"),
                ("First Bill Time", summary_data.get('first_bill_time', 'N/A')),
                ("Last Bill Time", summary_data.get('last_bill_time', 'N/A'))
            ]
            
            for label, value in summary_items:
                ws[f'A{current_row}'] = label
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            
            current_row += 2
            
            # === CATEGORY WISE SALES ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "CATEGORY WISE SALES"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            ws[f'A{current_row}'] = "Category"
            ws[f'B{current_row}'] = "Total Sales"
            ws[f'A{current_row}'].font = self.header_font
            ws[f'B{current_row}'].font = self.header_font
            ws[f'A{current_row}'].fill = self.header_fill
            ws[f'B{current_row}'].fill = self.header_fill
            ws[f'A{current_row}'].alignment = self.header_alignment
            ws[f'B{current_row}'].alignment = self.header_alignment
            current_row += 1
            
            category_totals = summary_data.get('category_totals', {})
            for category, total in category_totals.items():
                ws[f'A{current_row}'] = category.capitalize()
                ws[f'B{current_row}'] = f"{total:.2f}"
                ws[f'B{current_row}'].alignment = self.currency_alignment
                current_row += 1
            
            current_row += 2
            
            # === DETAILED BILLS ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "DETAILED BILLS"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            # Headers
            headers = ["Bill No", "Date", "Time", "Product ID", "Product Name", "Quantity", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            current_row += 1
            
            # Bill data
            for bill in bills:
                for item in bill.get('items', []):
                    ws[f'A{current_row}'] = bill.get('bill_no', '')
                    ws[f'B{current_row}'] = bill.get('created_at', '').split(' ')[0] if bill.get('created_at') else ''
                    ws[f'C{current_row}'] = bill.get('created_at', '').split(' ')[1] if bill.get('created_at') else ''
                    ws[f'D{current_row}'] = item.get('product_id', '')
                    ws[f'E{current_row}'] = item.get('name', '')
                    ws[f'F{current_row}'] = item.get('quantity', 0)
                    ws[f'G{current_row}'] = f"{item.get('price', 0) * item.get('quantity', 0):.2f}"
                    
                    # Apply formatting
                    for col in range(1, 8):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = self.data_font
                        cell.alignment = self.currency_alignment if col == 7 else self.data_alignment
                        cell.border = self.border
                    
                    current_row += 1
            
            # Save the workbook
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating detailed Excel report: {e}")
            return None
    
    def export_simple_sales_report(self, bills: List[Dict]) -> str:
        """Create a simple Excel report with basic bill details"""
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            filename = f"simple_sales_report_{today_str}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Simple Sales Report"
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            
            # Headers
            headers = ["Bill No", "Date", "Time", "Product ID", "Product Name", "Quantity", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            
            # Data
            current_row = 2
            for bill in bills:
                for item in bill.get('items', []):
                    ws[f'A{current_row}'] = bill.get('bill_no', '')
                    ws[f'B{current_row}'] = bill.get('created_at', '').split(' ')[0] if bill.get('created_at') else ''
                    ws[f'C{current_row}'] = bill.get('created_at', '').split(' ')[1] if bill.get('created_at') else ''
                    ws[f'D{current_row}'] = item.get('product_id', '')
                    ws[f'E{current_row}'] = item.get('name', '')
                    ws[f'F{current_row}'] = item.get('quantity', 0)
                    ws[f'G{current_row}'] = f"{item.get('price', 0) * item.get('quantity', 0):.2f}"
                    
                    # Apply formatting
                    for col in range(1, 8):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = self.data_font
                        cell.alignment = self.currency_alignment if col == 7 else self.data_alignment
                        cell.border = self.border
                    
                    current_row += 1
            
            # Save the workbook
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating simple Excel report: {e}")
            return None
    
    def export_summary_report(self, summary_data: Dict) -> str:
        """Create a summary Excel report"""
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            filename = f"summary_report_{today_str}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary Report"
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            
            # Title
            ws.merge_cells('A1:B1')
            ws['A1'] = "DAILY SALES SUMMARY"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            
            # Summary data
            summary_items = [
                ("Date", summary_data.get('date', today_str)),
                ("Total Bills", summary_data.get('total_bills', 0)),
                ("Total Sales", f"{summary_data.get('total_sales', 0):.2f}"),
                ("First Bill Time", summary_data.get('first_bill_time', 'N/A')),
                ("Last Bill Time", summary_data.get('last_bill_time', 'N/A'))
            ]
            
            current_row = 3
            for label, value in summary_items:
                ws[f'A{current_row}'] = label
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            
            current_row += 2
            
            # Category totals
            ws[f'A{current_row}'] = "CATEGORY WISE SALES"
            ws[f'A{current_row}'].font = Font(bold=True, size=14)
            current_row += 1
            
            ws[f'A{current_row}'] = "Category"
            ws[f'B{current_row}'] = "Total Sales"
            ws[f'A{current_row}'].font = self.header_font
            ws[f'B{current_row}'].font = self.header_font
            current_row += 1
            
            category_totals = summary_data.get('category_totals', {})
            for category, total in category_totals.items():
                ws[f'A{current_row}'] = category.capitalize()
                ws[f'B{current_row}'] = f"{total:.2f}"
                ws[f'B{current_row}'].alignment = self.currency_alignment
                current_row += 1
            
            # Save the workbook
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating summary Excel report: {e}")
            return None
    
    def create_sample_report(self) -> str:
        """Create a sample Excel report for demonstration when no bills exist"""
        try:
            today_str = date.today().strftime("%Y-%m-%d")
            filename = f"sample_sales_report_{today_str}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sample Sales Report"
            
            # Set column widths
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            current_row = 1
            
            # === SUMMARY SECTION ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "DAILY SALES SUMMARY (SAMPLE)"
            ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            # Sample summary data
            summary_items = [
                ("Date", today_str),
                ("Total Bills", 0),
                ("Total Sales", "0.00"),
                ("First Bill Time", "N/A"),
                ("Last Bill Time", "N/A")
            ]
            
            for label, value in summary_items:
                ws[f'A{current_row}'] = label
                ws[f'B{current_row}'] = value
                ws[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            
            current_row += 2
            
            # === CATEGORY WISE SALES ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "CATEGORY WISE SALES (SAMPLE)"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            ws[f'A{current_row}'] = "Category"
            ws[f'B{current_row}'] = "Total Sales"
            ws[f'A{current_row}'].font = self.header_font
            ws[f'B{current_row}'].font = self.header_font
            current_row += 1
            
            categories = ["coldrink", "paan", "other"]
            for category in categories:
                ws[f'A{current_row}'] = category.capitalize()
                ws[f'B{current_row}'] = "0.00"
                ws[f'B{current_row}'].alignment = self.currency_alignment
                current_row += 1
            
            current_row += 2
            
            # === DETAILED BILLS ===
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "DETAILED BILLS (SAMPLE DATA)"
            ws[f'A{current_row}'].font = Font(bold=True, size=14, color="FFFFFF")
            ws[f'A{current_row}'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            # Note
            ws.merge_cells(f'A{current_row}:G{current_row}')
            ws[f'A{current_row}'] = "Note: No bills found for today. This is a sample report format."
            ws[f'A{current_row}'].font = Font(italic=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal="center", vertical="center")
            current_row += 2
            
            # Headers
            headers = ["Bill No", "Date", "Time", "Product ID", "Product Name", "Quantity", "Total"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
            current_row += 1
            
            # Sample data
            sample_data = [
                ("SAMPLE001", today_str, "12:00:00", "SAMPLE001", "Sample Cold Drink", 2, 50.00),
                ("SAMPLE001", today_str, "12:00:00", "SAMPLE002", "Sample Paan", 1, 15.00)
            ]
            
            for bill_no, bill_date, time, product_id, name, quantity, total in sample_data:
                ws[f'A{current_row}'] = bill_no
                ws[f'B{current_row}'] = bill_date
                ws[f'C{current_row}'] = time
                ws[f'D{current_row}'] = product_id
                ws[f'E{current_row}'] = name
                ws[f'F{current_row}'] = quantity
                ws[f'G{current_row}'] = f"{total:.2f}"
                
                # Apply formatting
                for col in range(1, 8):
                    cell = ws.cell(row=current_row, column=col)
                    cell.font = self.data_font
                    cell.alignment = self.currency_alignment if col == 7 else self.data_alignment
                    cell.border = self.border
                
                current_row += 1
            
            # Save the workbook
            wb.save(filepath)
            return filepath
            
        except Exception as e:
            print(f"Error creating sample Excel report: {e}")
            return None
